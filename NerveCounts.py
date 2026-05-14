import os
import shutil
import pandas as pd
from tkinter import Tk, filedialog

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# -------------------------------------------------------
# Helper: Excel-safe sheet names (max 31 characters)
# -------------------------------------------------------
def safe_sheet_name(name, max_length=31):
    return name[:max_length]

# -------------------------------------------------------
# Formatting constants (Processed Data only)
# -------------------------------------------------------
NO_BORDER = Border(
    left=Side(border_style=None),
    right=Side(border_style=None),
    top=Side(border_style=None),
    bottom=Side(border_style=None)
)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_FONT = Font(bold=True)

# -------------------------------------------------------
# Format ONLY the Processed Data sheet
# -------------------------------------------------------
def format_excel_file(file_path, sheet_name="Processed Data"):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Format header row
    for cell in ws[1]:
        cell.alignment = HEADER_ALIGNMENT
        cell.font = HEADER_FONT
        cell.border = NO_BORDER

    # Autosize all columns
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)

# -------------------------------------------------------
# Add shifted coordinates, angles, quadrants, nerve count
# -------------------------------------------------------
def add_shifted_coordinates(file_path, sheet_name="Processed Data"):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    col_letters = {cell.value: cell.column_letter for cell in ws[1]}
    required_cols = [
        'Centre X lumen (µm)', 'Centre Y lumen (µm)',
        'Centre X nerve bundle (µm)', 'Centre Y nerve bundle (µm)'
    ]
    for col in required_cols:
        if col not in col_letters:
            raise ValueError(f"Column '{col}' not found in {file_path}")

    lumen_x_cell = f"${col_letters['Centre X lumen (µm)']}$2"
    lumen_y_cell = f"${col_letters['Centre Y lumen (µm)']}$2"

    max_col = ws.max_column
    shifted_x_col = max_col + 1
    shifted_y_col = max_col + 2
    angle_rad_col = max_col + 3
    angle_deg_col = max_col + 4
    quadrant_col  = max_col + 5

    headers = [
        ("Shifted X (µm)", shifted_x_col),
        ("Shifted Y (µm)", shifted_y_col),
        ("Angle between Nerve Bundle and Lumen (rad)", angle_rad_col),
        ("Angle between Nerve Bundle and Lumen (deg)", angle_deg_col),
        ("Quadrant", quadrant_col)
    ]
    for header, col in headers:
        h = ws.cell(row=1, column=col, value=header)
        h.alignment = HEADER_ALIGNMENT
        h.font = HEADER_FONT
        h.border = NO_BORDER

    for row in range(2, ws.max_row + 1):
        ws.cell(
            row=row,
            column=shifted_x_col,
            value=f"={col_letters['Centre X nerve bundle (µm)']}{row}-{lumen_x_cell}"
        )
        ws.cell(
            row=row,
            column=shifted_y_col,
            value=f"={col_letters['Centre Y nerve bundle (µm)']}{row}-{lumen_y_cell}"
        )

        sx = get_column_letter(shifted_x_col)
        sy = get_column_letter(shifted_y_col)
        ar = get_column_letter(angle_rad_col)
        ad = get_column_letter(angle_deg_col)

        ws.cell(row=row, column=angle_rad_col, value=f"=ATAN2({sx}{row},{sy}{row})")
        ws.cell(
            row=row,
            column=angle_deg_col,
            value=f"=IF({ar}{row}*180/PI()<0,{ar}{row}*180/PI()+360,{ar}{row}*180/PI())"
        )

        ws.cell(
            row=row,
            column=quadrant_col,
            value=(
                f'=IF(AND({ad}{row}>=45,{ad}{row}<135),"Quadrant 1",'
                f'IF(AND({ad}{row}>=135,{ad}{row}<225),"Quadrant 4",'
                f'IF(AND({ad}{row}>=225,{ad}{row}<315),"Quadrant 3","Quadrant 2")))'
            )
        )

    quadrant_letter = get_column_letter(quadrant_col)

    # --------------------- Nerve count summary ---------------------
    bin_col = quadrant_col + 1
    count_col = quadrant_col + 2

    summary_headers = ["Quadrant Bin", "Number per Quadrant Bin"]
    for i, h in enumerate(summary_headers, start=bin_col):
        cell = ws.cell(row=1, column=i, value=h)
        cell.alignment = HEADER_ALIGNMENT
        cell.font = HEADER_FONT
        cell.border = NO_BORDER

    quadrants = ["Quadrant 1", "Quadrant 4", "Quadrant 3", "Quadrant 2"]
    for i, q in enumerate(quadrants, start=2):
        ws.cell(row=i, column=bin_col, value=q)
        ws.cell(
            row=i,
            column=count_col,
            value=f"=COUNTIF({quadrant_letter}2:{quadrant_letter}{ws.max_row},{get_column_letter(bin_col)}{i})"
        )

    total_row = 2 + len(quadrants)
    ws.cell(row=total_row, column=bin_col, value="Total")
    ws.cell(
        row=total_row,
        column=count_col,
        value=f"=SUM({get_column_letter(count_col)}2:{get_column_letter(count_col)}{total_row-1})"
    )

    wb.save(file_path)

# -------------------------------------------------------
# Main pipeline
# -------------------------------------------------------
def select_and_process_tsv_files():
    root = Tk()
    root.withdraw()

    file_paths = filedialog.askopenfilenames(
        title="Select TSV Files",
        filetypes=[("TSV Files", "*.tsv")]
    )

    if not file_paths:
        return

    parent_dir = os.path.dirname(file_paths[0])
    output_folder = os.path.join(parent_dir, "TSV and Excel files")
    os.makedirs(output_folder, exist_ok=True)

    for file_path in file_paths:
        file_name = os.path.basename(file_path)
        base = os.path.splitext(file_name)[0]

        new_tsv_path = os.path.join(output_folder, file_name)
        shutil.move(file_path, new_tsv_path)

        raw_df = pd.read_csv(new_tsv_path, sep="\t")
        processed_df = raw_df.copy()

        processed_df.drop(
            columns=[c for c in [
                "Study level 1", "Study level 2", "Study level 3",
                "Image", "LayerData"
            ] if c in processed_df.columns],
            inplace=True
        )

        processed_df.rename(columns={
            "Centre X lumen": "Centre X lumen (µm)",
            "Centre Y lumen": "Centre Y lumen (µm)",
            "Centre X nerve bundle": "Centre X nerve bundle (µm)",
            "Centre Y nerve bundle": "Centre Y nerve bundle (µm)",
            "Nerve Bundle Distance": "Nerve Bundle Distance (µm)"
        }, inplace=True)

        excel_path = os.path.join(output_folder, base + ".xlsx")

        sheet_name = safe_sheet_name(base)

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name=sheet_name, index=False)
            processed_df.to_excel(writer, sheet_name="Processed Data", index=False)

        wb = load_workbook(excel_path)
        raw_ws = wb[sheet_name]
        for cell in raw_ws[1]:
            cell.font = Font(bold=False)
            cell.border = NO_BORDER
        wb.save(excel_path)

        format_excel_file(excel_path, "Processed Data")
        add_shifted_coordinates(excel_path, "Processed Data")

        print(f"Processed {base}")

    print("All files processed successfully.")

if __name__ == "__main__":
    select_and_process_tsv_files()
